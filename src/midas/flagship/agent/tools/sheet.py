"""Spreadsheet tools — the wow-moment from PDF to Excel, gated + receipted.

``sheet.read`` is AUTO (``read_local_files``). ``sheet.write`` is APPROVE
(``write_spreadsheet``): the Toolset queues an approval card carrying the exact
cell range + sha256 before; the actual write happens via
:func:`execute_sheet_write` after the human resolves the approval.

Privacy: receipts record the cell *address range* + count + sha256 before/after,
NOT the raw values. The data is in the workspace file the operator already owns;
the chain proves *what was written where*, not what the value was.

Dependencies are lazy: openpyxl is required for .xlsx and lives behind the
``[sheets]`` extra. CSV needs no extra dep.
"""

from __future__ import annotations

import csv
import hashlib
import io
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from .fsguard import FsGuard

CellAddr = str  # e.g. "A1", "B14"
CellValue = Any  # primitives only; openpyxl rejects arbitrary objects


@dataclass(frozen=True)
class SheetReadResult:
    path: str
    sheet_name: str
    rows: list[list[CellValue]]
    n_rows: int
    n_cols: int


@dataclass
class SheetWritePlan:
    """Approval payload for a gated spreadsheet write."""

    path: str
    sheet_name: str
    cells: list[tuple[CellAddr, CellValue]]
    sha256_prev: str | None = None
    notes: list[str] = field(default_factory=list)

    @property
    def cell_count(self) -> int:
        return len(self.cells)

    @property
    def cell_range(self) -> str:
        if not self.cells:
            return ""
        addrs = [c[0] for c in self.cells]
        # Min/max by row + col, return a tight A1:B7-style range.
        rows = [_parse_addr(a)[1] for a in addrs]
        cols = [_parse_addr(a)[0] for a in addrs]
        return f"{_addr(min(cols), min(rows))}:{_addr(max(cols), max(rows))}"


def _sha256_file(path: Path) -> str | None:
    if not path.exists() or not path.is_file():
        return None
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _is_csv(path: Path) -> bool:
    return path.suffix.lower() == ".csv"


def _is_xlsx(path: Path) -> bool:
    return path.suffix.lower() in {".xlsx", ".xlsm"}


# ── sheet.read ────────────────────────────────────────────────────────────────


def sheet_read(
    guard: FsGuard, path: str, *, sheet_name: str | None = None, max_rows: int = 5000
) -> SheetReadResult:
    target = guard.resolve(path, must_exist=True)
    if _is_csv(target):
        with target.open("r", encoding="utf-8", newline="") as f:
            rows = list(csv.reader(f))[:max_rows]
        return SheetReadResult(
            path=str(target),
            sheet_name=sheet_name or target.stem,
            rows=[[c for c in r] for r in rows],
            n_rows=len(rows),
            n_cols=max((len(r) for r in rows), default=0),
        )
    if _is_xlsx(target):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - exercised by the test path
            raise RuntimeError(
                "sheet.read for .xlsx requires the [sheets] extra (openpyxl)"
            ) from exc
        wb = load_workbook(target, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append(list(row))
        return SheetReadResult(
            path=str(target),
            sheet_name=str(ws.title),
            rows=rows,
            n_rows=len(rows),
            n_cols=max((len(r) for r in rows), default=0),
        )
    raise ValueError(f"unsupported spreadsheet extension: {target.suffix!r}")


# ── sheet.write — plan-only, no mutation ─────────────────────────────────────


def plan_sheet_write(
    guard: FsGuard,
    path: str,
    *,
    sheet_name: str = "Sheet1",
    cells: list[tuple[CellAddr, CellValue]],
) -> SheetWritePlan:
    """Build the approval payload for a sheet.write. Does NOT mutate the file."""
    target = guard.resolve(path)
    if not (_is_csv(target) or _is_xlsx(target)):
        raise ValueError(f"unsupported spreadsheet extension: {target.suffix!r}")
    if not cells:
        raise ValueError("plan_sheet_write requires at least one cell")
    for addr, _ in cells:
        _parse_addr(addr)  # raises if shape is bad
    return SheetWritePlan(
        path=str(target),
        sheet_name=sheet_name,
        cells=list(cells),
        sha256_prev=_sha256_file(target),
    )


# ── sheet.write — executed after approval ─────────────────────────────────────


def execute_sheet_write(guard: FsGuard, plan: SheetWritePlan) -> dict[str, Any]:
    """Apply the plan to disk. Caller MUST have confirmed approval already."""
    target = guard.resolve(plan.path)
    target.parent.mkdir(parents=True, exist_ok=True)
    sha_prev = _sha256_file(target)
    if _is_csv(target):
        cells_written = _write_csv_cells(target, plan)
    elif _is_xlsx(target):
        cells_written = _write_xlsx_cells(target, plan)
    else:
        raise ValueError(f"unsupported spreadsheet extension: {target.suffix!r}")
    sha_new = _sha256_file(target) or ""
    return {
        "path": str(target),
        "sheet": plan.sheet_name,
        "cells_written": cells_written,
        "cell_range": plan.cell_range,
        "sha256_prev": sha_prev,
        "sha256_new": sha_new,
    }


def _write_csv_cells(target: Path, plan: SheetWritePlan) -> int:
    # Load existing matrix (if any), expand to fit, write requested cells, save.
    rows: list[list[str]] = []
    if target.exists():
        with target.open("r", encoding="utf-8", newline="") as f:
            rows = list(csv.reader(f))
    for addr, val in plan.cells:
        col, row = _parse_addr(addr)
        while len(rows) < row:
            rows.append([])
        line = rows[row - 1]
        while len(line) < col:
            line.append("")
        line[col - 1] = "" if val is None else str(val)
    buf = io.StringIO()
    csv.writer(buf).writerows(rows)
    target.write_text(buf.getvalue(), encoding="utf-8")
    return len(plan.cells)


def _write_xlsx_cells(target: Path, plan: SheetWritePlan) -> int:
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "sheet.write for .xlsx requires the [sheets] extra (openpyxl)"
        ) from exc
    if target.exists():
        wb = load_workbook(target)
        ws = wb[plan.sheet_name] if plan.sheet_name in wb.sheetnames else wb.create_sheet(
            plan.sheet_name
        )
    else:
        wb = Workbook()
        # `Workbook()` creates a default sheet named "Sheet"; rename or add.
        if plan.sheet_name and plan.sheet_name != wb.active.title:
            wb.active.title = plan.sheet_name
        ws = wb.active
    for addr, val in plan.cells:
        ws[addr] = val
    wb.save(target)
    return len(plan.cells)


# ── address helpers ──────────────────────────────────────────────────────────


_ADDR_RE = re.compile(r"^([A-Z]+)([1-9][0-9]*)$")


def _parse_addr(addr: CellAddr) -> tuple[int, int]:
    """Parse an A1-style address → (column 1-based, row 1-based)."""
    if not isinstance(addr, str):
        raise ValueError(f"cell address must be a string, got {type(addr).__name__}")
    m = _ADDR_RE.match(addr.upper())
    if not m:
        raise ValueError(f"invalid cell address: {addr!r}")
    col_letters, row_str = m.groups()
    col = 0
    for ch in col_letters:
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return col, int(row_str)


def _addr(col: int, row: int) -> CellAddr:
    letters = ""
    n = col
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters = chr(ord("A") + rem) + letters
    return f"{letters}{row}"
